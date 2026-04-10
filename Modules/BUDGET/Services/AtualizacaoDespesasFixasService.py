import os
import re
import shutil
import uuid
from datetime import date, datetime

import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from werkzeug.utils import secure_filename


class AtualizacaoDespesasFixasService:
    PASTA_TEMPORARIA = os.path.abspath(
        os.path.join(os.path.dirname(__file__), '../../..', 'Data', 'Temp', 'BudgetAtualizacaoDespesasFixas')
    )

    EXTENSOES_ORIGEM_PERMITIDAS = {'.xlsx', '.xls'}
    EXTENSOES_DESTINO_PERMITIDAS = {'.xlsm', '.xlsx'}

    COLUNAS_IDENTIFICADORAS = [
        'Empresa',
        'Conta Contábil',
        'Descrição',
        'Favorecido',
        'Centro de Custos',
        'Filial',
    ]

    COLUNAS_DESTINO = [
        'Tabela',
        'Departamento',
        'ID - Formulário',
        'Conta Contábil - Fórmula',
        'Código Conta',
        'Empresa',
        'Conta Contábil',
        'Descrição',
        'Favorecido',
        'Centro de Custos',
        'Item',
        'Filial',
        'Cliente (Item Conta)',
        'Valor',
        'Mês',
    ]

    COLUNA_DEPARTAMENTO_ORIGEM = '__departamento_origem__'

    MESES_PT_BR = {
        1: 'Janeiro',
        2: 'Fevereiro',
        3: 'Março',
        4: 'Abril',
        5: 'Maio',
        6: 'Junho',
        7: 'Julho',
        8: 'Agosto',
        9: 'Setembro',
        10: 'Outubro',
        11: 'Novembro',
        12: 'Dezembro',
    }

    REGEX_COLUNA_DATA = re.compile(r'^(\d{4}[-/]\d{2}([-/]\d{2})?|\d{2}[-/]\d{2}[-/]\d{4})(\s.+)?$')

    def __init__(self):
        self._garantir_pasta_temporaria()

    def salvarArquivoOrigem(self, arquivo_storage):
        return self._salvarArquivoTemporario(arquivo_storage, self.EXTENSOES_ORIGEM_PERMITIDAS, 'origem')

    def salvarArquivoDestino(self, arquivo_storage):
        return self._salvarArquivoTemporario(arquivo_storage, self.EXTENSOES_DESTINO_PERMITIDAS, 'destino')

    def listarAbasDestino(self, token_arquivo_destino):
        caminho_arquivo = self._resolverCaminhoTemporario(token_arquivo_destino, prefixo='destino_')
        workbook = openpyxl.load_workbook(caminho_arquivo, read_only=True, keep_vba=True)
        try:
            return workbook.sheetnames
        finally:
            workbook.close()

    def processarAtualizacao(self, token_arquivo_origem, token_arquivo_destino, nome_aba_destino):
        caminho_origem = self._resolverCaminhoTemporario(token_arquivo_origem, prefixo='origem_')
        caminho_destino_original = self._resolverCaminhoTemporario(token_arquivo_destino, prefixo='destino_')

        if not nome_aba_destino or not str(nome_aba_destino).strip():
            raise ValueError('Selecione uma aba de destino para continuar.')

        nome_aba_destino = str(nome_aba_destino).strip()
        dataframe_final = self._montarDataFrameFinal(caminho_origem)

        nome_base_destino, extensao_destino = os.path.splitext(os.path.basename(caminho_destino_original))
        timestamp_sufixo = datetime.now().strftime('%Y%m%d_%H%M%S')
        nome_arquivo_saida = f"resultado_{uuid.uuid4().hex[:8]}_{nome_base_destino}_Atualizado_{timestamp_sufixo}{extensao_destino}"
        caminho_arquivo_saida = os.path.join(self.PASTA_TEMPORARIA, nome_arquivo_saida)

        shutil.copy2(caminho_destino_original, caminho_arquivo_saida)

        workbook_injecao = openpyxl.load_workbook(caminho_arquivo_saida, keep_vba=True)
        try:
            if nome_aba_destino not in workbook_injecao.sheetnames:
                raise ValueError(f"A aba '{nome_aba_destino}' não existe no arquivo de destino informado.")

            planilha_alvo = workbook_injecao[nome_aba_destino]
            ultima_linha_real = self._encontrarUltimaLinhaValida(planilha_alvo, indice_coluna_base=9)
            linha_vazia_disponivel = ultima_linha_real + 1

            for linha_dados in dataframe_to_rows(dataframe_final, index=False, header=False):
                for indice_coluna, valor_celula in enumerate(linha_dados, start=4):
                    if pd.notna(valor_celula):
                        planilha_alvo.cell(row=linha_vazia_disponivel, column=indice_coluna, value=valor_celula)
                linha_vazia_disponivel += 1

            workbook_injecao.save(caminho_arquivo_saida)
        finally:
            workbook_injecao.close()

        return {
            'tokenDownload': nome_arquivo_saida,
            'nomeArquivo': nome_arquivo_saida,
            'linhasInseridas': len(dataframe_final.index),
            'abaDestino': nome_aba_destino,
        }

    def obterCaminhoArquivoProcessado(self, token_download):
        return self._resolverCaminhoTemporario(token_download, prefixo='resultado_')

    def _garantir_pasta_temporaria(self):
        os.makedirs(self.PASTA_TEMPORARIA, exist_ok=True)

    def _salvarArquivoTemporario(self, arquivo_storage, extensoes_permitidas, prefixo):
        if arquivo_storage is None or not getattr(arquivo_storage, 'filename', ''):
            raise ValueError('Nenhum arquivo foi enviado.')

        nome_original = secure_filename(arquivo_storage.filename)
        extensao = os.path.splitext(nome_original)[1].lower()
        if extensao not in extensoes_permitidas:
            extensoes_texto = ', '.join(sorted(extensoes_permitidas))
            raise ValueError(f'Formato inválido. Utilize um arquivo {extensoes_texto}.')

        nome_unico = f"{prefixo}_{uuid.uuid4().hex}_{nome_original}"
        caminho_arquivo = os.path.join(self.PASTA_TEMPORARIA, nome_unico)
        arquivo_storage.save(caminho_arquivo)

        return {
            'token': nome_unico,
            'nomeArquivo': nome_original,
        }

    def _resolverCaminhoTemporario(self, token_arquivo, prefixo=None):
        token_seguro = os.path.basename(str(token_arquivo or '').strip())
        if not token_seguro:
            raise ValueError('Arquivo temporário inválido ou expirado.')

        if prefixo and not token_seguro.startswith(prefixo):
            raise ValueError('Tipo de arquivo temporário inválido para esta operação.')

        caminho_arquivo = os.path.join(self.PASTA_TEMPORARIA, token_seguro)
        if not os.path.exists(caminho_arquivo):
            raise ValueError('Arquivo temporário inválido ou expirado.')
        return caminho_arquivo

    def _montarDataFrameFinal(self, caminho_origem):
        dataframe_origem = pd.read_excel(caminho_origem, sheet_name=0, header=3, engine='openpyxl')
        dataframe_origem.columns = [self._normalizarCabecalho(coluna) for coluna in dataframe_origem.columns]
        serie_departamento = self._extrairColunaDepartamento(dataframe_origem)

        colunas_obrigatorias = [
            coluna for coluna in self.COLUNAS_IDENTIFICADORAS if coluna not in dataframe_origem.columns
        ]
        if colunas_obrigatorias:
            raise ValueError(
                'A planilha base não contém todas as colunas obrigatórias. '
                f"Faltando: {', '.join(colunas_obrigatorias)}."
            )

        mapa_colunas_meses = self._identificarColunasMeses(dataframe_origem.columns)
        if not mapa_colunas_meses:
            raise ValueError('Nenhuma coluna mensal foi identificada na planilha base. Valide o cabeçalho do arquivo.')

        colunas_meses = list(mapa_colunas_meses.keys())
        dataframe_filtrado = dataframe_origem[self.COLUNAS_IDENTIFICADORAS + colunas_meses].copy()
        dataframe_filtrado.insert(0, self.COLUNA_DEPARTAMENTO_ORIGEM, serie_departamento.values)

        dataframe_transformado = pd.melt(
            dataframe_filtrado,
            id_vars=[self.COLUNA_DEPARTAMENTO_ORIGEM] + self.COLUNAS_IDENTIFICADORAS,
            value_vars=colunas_meses,
            var_name='Mês',
            value_name='Valor',
        )

        dataframe_transformado['Valor'] = pd.to_numeric(dataframe_transformado['Valor'], errors='coerce')
        dataframe_transformado = dataframe_transformado.dropna(subset=['Valor'])
        dataframe_transformado = dataframe_transformado[dataframe_transformado['Valor'] != 0]

        if dataframe_transformado.empty:
            raise ValueError('Nenhum valor de budget válido foi encontrado para importar.')

        dataframe_transformado['Empresa'] = dataframe_transformado['Empresa'].astype(str).str.strip().str.capitalize()
        dataframe_transformado['Mês'] = dataframe_transformado['Mês'].map(mapa_colunas_meses)
        dataframe_transformado = dataframe_transformado.dropna(subset=['Mês'])

        dataframe_final = pd.DataFrame(columns=self.COLUNAS_DESTINO)
        dataframe_final['Departamento'] = dataframe_transformado[self.COLUNA_DEPARTAMENTO_ORIGEM]
        dataframe_final['Empresa'] = dataframe_transformado['Empresa']
        dataframe_final['Conta Contábil'] = dataframe_transformado['Conta Contábil']
        dataframe_final['Descrição'] = dataframe_transformado['Descrição']
        dataframe_final['Favorecido'] = dataframe_transformado['Favorecido']
        dataframe_final['Centro de Custos'] = dataframe_transformado['Centro de Custos']
        dataframe_final['Filial'] = dataframe_transformado['Filial']
        dataframe_final['Valor'] = dataframe_transformado['Valor']
        dataframe_final['Mês'] = dataframe_transformado['Mês']
        dataframe_final['Conta Contábil - Fórmula'] = dataframe_transformado['Conta Contábil']
        dataframe_final['Tabela'] = 'Budget'
        dataframe_final['ID - Formulário'] = None
        dataframe_final['Código Conta'] = None
        dataframe_final['Item'] = None
        dataframe_final['Cliente (Item Conta)'] = None

        return dataframe_final[self.COLUNAS_DESTINO]

    def _extrairColunaDepartamento(self, dataframe_origem):
        if dataframe_origem.shape[1] == 0:
            raise ValueError('A planilha base não possui dados suficientes para identificar o departamento na coluna A.')

        serie_departamento = dataframe_origem.iloc[:, 0].copy()
        return serie_departamento.apply(self._normalizarTextoOpcional)

    def _identificarColunasMeses(self, colunas):
        mapa_meses = {}
        for coluna in colunas:
            data_referencia = self._extrairDataCabecalho(coluna)
            if data_referencia is None:
                continue
            mapa_meses[coluna] = self.MESES_PT_BR.get(data_referencia.month)
        return mapa_meses

    def _extrairDataCabecalho(self, cabecalho):
        if isinstance(cabecalho, (datetime, date, pd.Timestamp)):
            data_referencia = pd.to_datetime(cabecalho, errors='coerce')
            return None if pd.isna(data_referencia) else data_referencia

        texto = str(cabecalho or '').strip()
        if not texto or not self.REGEX_COLUNA_DATA.match(texto):
            return None

        data_referencia = pd.to_datetime(texto, errors='coerce')
        if pd.isna(data_referencia):
            return None
        return data_referencia

    def _normalizarCabecalho(self, valor_cabecalho):
        if isinstance(valor_cabecalho, (datetime, date, pd.Timestamp)):
            return pd.to_datetime(valor_cabecalho).strftime('%Y-%m-%d')
        return str(valor_cabecalho).strip()

    def _normalizarTextoOpcional(self, valor):
        if pd.isna(valor):
            return None

        texto = str(valor).strip()
        return texto or None

    def _encontrarUltimaLinhaValida(self, planilha, indice_coluna_base=9):
        for linha_atual in range(planilha.max_row, 0, -1):
            if planilha.cell(row=linha_atual, column=indice_coluna_base).value is not None:
                return linha_atual
        return 1