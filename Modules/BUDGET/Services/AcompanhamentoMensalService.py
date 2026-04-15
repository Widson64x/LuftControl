import os
import uuid
from copy import copy
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import extract, func

from Db.Connections import GetSqlServerSession
from Models.SqlServer.Budget import Budget, BudgetItem
from Models.SqlServer.ContaPagar import ContaPagar, ContaPagarNotaFiscal, PlanoConta
from Models.SqlServer.Fornecedor import Fornecedor
from Modules.SISTEMA.Services.GestorCentroCustoService import GestorCentroCustoService


class AcompanhamentoMensalService:
    PASTA_BASE = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../..', 'Data', 'Base'))
    ARQUIVO_TEMPLATE = os.path.join(PASTA_BASE, 'Template - budget - MENSAL.xlsx')
    PASTA_TEMPORARIA = os.path.abspath(
        os.path.join(os.path.dirname(__file__), '../../..', 'Data', 'Temp', 'BudgetAcompanhamentoMensal')
    )

    STATUS_EM_APROVACAO = (1, 2)
    STATUS_APROVADO = (3, 5)
    STATUS_CONSIDERADOS = STATUS_EM_APROVACAO + STATUS_APROVADO
    LINHA_INTRO_DETALHE = 18
    LINHA_INICIAL_DETALHE = 19
    QUANTIDADE_MINIMA_LINHAS = 9
    COLUNAS_PLANILHA = tuple(range(1, 11))
    FORMATO_DATA = 'dd/mm/yyyy'
    FORMATO_MOEDA = '#,##0.00'

    MESES = (
        (1, 'Janeiro', 'Jan', BudgetItem.Valor_JaneiroO),
        (2, 'Fevereiro', 'Fev', BudgetItem.Valor_FevereiroO),
        (3, 'Março', 'Mar', BudgetItem.Valor_MarcoO),
        (4, 'Abril', 'Abr', BudgetItem.Valor_AbrilO),
        (5, 'Maio', 'Mai', BudgetItem.Valor_MaioO),
        (6, 'Junho', 'Jun', BudgetItem.Valor_JunhoO),
        (7, 'Julho', 'Jul', BudgetItem.Valor_JulhoO),
        (8, 'Agosto', 'Ago', BudgetItem.Valor_AgostoO),
        (9, 'Setembro', 'Set', BudgetItem.Valor_SetembroO),
        (10, 'Outubro', 'Out', BudgetItem.Valor_OutubroO),
        (11, 'Novembro', 'Nov', BudgetItem.Valor_NovembroO),
        (12, 'Dezembro', 'Dez', BudgetItem.Valor_DezembroO),
    )

    def __init__(self):
        self._gestor_service = GestorCentroCustoService()
        self._garantirEstrutura()

    def obterContextoGestor(self, codigo_usuario):
        gestor = self._gestor_service.obterGestorConfigurado(codigo_usuario)
        agora = datetime.now()

        return {
            'gestor': gestor,
            'centros_custo': list(gestor.get('centros_custo', [])) if gestor else [],
            'ano_padrao': agora.year,
            'ano_opcoes': [agora.year - indice for indice in range(5)],
            'meses_preenchidos': agora.month,
            'template_disponivel': os.path.exists(self.ARQUIVO_TEMPLATE),
        }

    def gerarArquivo(self, codigo_usuario, ano=None, codigo_centro_custo=None, codigos_conta_contabil=None):
        agora = datetime.now()
        ano_referencia = self._normalizarAno(ano or agora.year, agora.year)
        gestor = self._obterGestorObrigatorio(codigo_usuario)
        centro = self._resolverCentroSelecionado(gestor, codigo_centro_custo)
        contas_contabeis = self._normalizarContasContabeis(codigos_conta_contabil)

        sessao = GetSqlServerSession()
        try:
            codigo_centro_decimal = self._converterCodigoDecimal(centro['codigo'])
            totais_budget = self._obterTotaisBudget(
                sessao,
                ano_referencia,
                codigo_centro_decimal,
                contas_contabeis,
            )
            lancamentos_mensais = self._obterLancamentosMensais(
                sessao,
                ano_referencia,
                codigo_centro_decimal,
                contas_contabeis,
            )
        finally:
            sessao.close()

        workbook = load_workbook(self.ARQUIVO_TEMPLATE)
        try:
            self._atualizarAbaLista(workbook['Lista'], ano_referencia, centro)

            linha_acumulada_anterior = None
            nome_aba_anterior = None
            for numero_mes, nome_aba, abreviacao, _ in self.MESES:
                planilha = workbook[nome_aba]
                lancamentos_mes = lancamentos_mensais.get(numero_mes, [])
                linha_acumulada_atual = self._preencherAbaMensal(
                    planilha=planilha,
                    ano=ano_referencia,
                    numero_mes=numero_mes,
                    abreviacao_mes=abreviacao,
                    gestor=gestor,
                    centro=centro,
                    total_ano=totais_budget['total_ano'],
                    total_mes=totais_budget['meses'].get(numero_mes, 0.0),
                    lancamentos=lancamentos_mes,
                    nome_aba_anterior=nome_aba_anterior,
                    linha_acumulada_anterior=linha_acumulada_anterior,
                )
                nome_aba_anterior = nome_aba
                linha_acumulada_anterior = linha_acumulada_atual

            nome_seguro_centro = self._normalizarNomeArquivo(centro['numero'])
            nome_seguro_responsavel = self._normalizarNomeArquivo(gestor['nome_usuario'])
            nome_arquivo = (
                f"acompanhamento_{uuid.uuid4().hex[:8]}_{ano_referencia}_{nome_seguro_centro}_{nome_seguro_responsavel}.xlsx"
            )
            caminho_saida = os.path.join(self.PASTA_TEMPORARIA, nome_arquivo)
            workbook.save(caminho_saida)
        finally:
            workbook.close()

        return {
            'tokenDownload': nome_arquivo,
            'nomeArquivo': nome_arquivo,
            'ano': ano_referencia,
            'centroCusto': centro,
            'gestor': {
                'codigo_usuario': gestor['codigo_usuario'],
                'nome_usuario': gestor['nome_usuario'],
                'cargo': gestor.get('cargo') or 'Gestor',
            },
            'mesesPreenchidos': min(agora.month, 12) if ano_referencia == agora.year else 12,
        }

    def obterCaminhoArquivoGerado(self, token_download):
        token_seguro = os.path.basename(str(token_download or '').strip())
        if not token_seguro or not token_seguro.startswith('acompanhamento_'):
            raise ValueError('Arquivo gerado inválido ou expirado.')

        caminho_arquivo = os.path.join(self.PASTA_TEMPORARIA, token_seguro)
        if not os.path.exists(caminho_arquivo):
            raise ValueError('Arquivo gerado inválido ou expirado.')

        return caminho_arquivo

    def _garantirEstrutura(self):
        os.makedirs(self.PASTA_TEMPORARIA, exist_ok=True)
        if not os.path.exists(self.ARQUIVO_TEMPLATE):
            raise ValueError(
                'O template padrão do Budget não foi encontrado em Data/Base/Template - budget - MENSAL.xlsx.'
            )

    def _normalizarAno(self, ano, ano_atual):
        try:
            ano_normalizado = int(str(ano).strip())
        except (TypeError, ValueError):
            raise ValueError('Ano inválido para gerar a planilha de acompanhamento.')

        if ano_normalizado < 2020 or ano_normalizado > ano_atual:
            raise ValueError('Selecione um ano atual ou anterior para gerar a planilha.')

        return ano_normalizado

    def _obterGestorObrigatorio(self, codigo_usuario):
        gestor = self._gestor_service.obterGestorConfigurado(codigo_usuario)
        if not gestor:
            raise ValueError('Seu usuário não possui centros de custo configurados para essa rotina.')

        centros = gestor.get('centros_custo', [])
        if not centros:
            raise ValueError('Seu usuário não possui centros de custo configurados para essa rotina.')

        return gestor

    def _resolverCentroSelecionado(self, gestor, codigo_centro_custo):
        # Cria uma cópia da lista de centros para evitar mutações acidentais e 
        # facilitar a busca
        centros = list(gestor.get('centros_custo', []))
        
        # Se a lista tiver apenas um centro e nenhum código for fornecido, 
        # retorna esse centro diretamente
        if len(centros) == 1 and not codigo_centro_custo:
            return centros[0]

        # Cria uma versão normalizada do código de centro de custo para comparação,
        codigo_normalizado = str(codigo_centro_custo or '').strip()
        # Se o código normalizado estiver vazio, significa que o usuário não 
        # selecionou um centro válido
        if not codigo_normalizado:
            raise ValueError('Selecione o centro de custo que sera usado para gerar a planilha.')
        
        # Para para cada centro de custo na lista do gestor, normaliza o código do centro e 
        # compara com o código fornecido
        for centro in centros:
            if str(centro.get('codigo')) == codigo_normalizado:
                # Retorna o centro correspondente se encontrar uma correspondência exata
                return centro

        raise ValueError('O centro de custo selecionado nao pertence ao gestor informado.')

    def _converterCodigoDecimal(self, valor):
        try:
            return Decimal(str(valor).strip())
        except (InvalidOperation, AttributeError, ValueError):
            raise ValueError('Codigo de centro de custo invalido para gerar a planilha.')

    def _normalizarContasContabeis(self, codigos_conta_contabil):
        if codigos_conta_contabil is None:
            return None

        if isinstance(codigos_conta_contabil, (list, tuple, set)):
            valores = list(codigos_conta_contabil)
        else:
            texto = str(codigos_conta_contabil).strip()
            if not texto or texto.upper() == 'TODOS':
                return None
            valores = texto.split(',')

        contas = []
        contas_vistas = set()

        for valor in valores:
            texto = str(valor or '').strip()
            if not texto:
                continue

            if texto.upper() == 'TODOS':
                return None

            try:
                conta_decimal = self._converterCodigoDecimal(texto)
            except ValueError:
                raise ValueError('Codigo de conta contabil invalido para gerar a planilha.')

            chave = str(conta_decimal)
            if chave in contas_vistas:
                continue

            contas_vistas.add(chave)
            contas.append(conta_decimal)

        return contas or None

    def _obterTotaisBudget(self, sessao, ano, codigo_centro_custo, codigos_conta_contabil=None):
        colunas_mes = [
            func.coalesce(func.sum(coluna), 0).label(f'mes_{numero}')
            for numero, _, _, coluna in self.MESES
        ]

        query = (
            sessao.query(*colunas_mes)
            .select_from(BudgetItem)
            .join(Budget, BudgetItem.Codigo_Budget == Budget.Codigo_Budget)
            .filter(Budget.Ano_Vigencia == ano)
            .filter(BudgetItem.Codigo_CentroCusto == codigo_centro_custo)
        )

        if codigos_conta_contabil:
            query = query.filter(BudgetItem.Codigo_ContaContabil.in_(codigos_conta_contabil))

        registro = query.one()

        meses = {
            numero: float(getattr(registro, f'mes_{numero}') or 0.0)
            for numero, _, _, _ in self.MESES
        }
        return {
            'meses': meses,
            'total_ano': sum(meses.values()),
        }

    def _obterLancamentosMensais(self, sessao, ano, codigo_centro_custo, codigos_conta_contabil=None):
        subconsulta_data_nota = (
            sessao.query(
                ContaPagarNotaFiscal.Codigo_ContaPagar.label('codigoContaPagar'),
                func.max(ContaPagarNotaFiscal.Data_Digitacao).label('dataDigitacaoNotaFiscal'),
            )
            .group_by(ContaPagarNotaFiscal.Codigo_ContaPagar)
            .subquery()
        )

        data_efetiva = func.coalesce(subconsulta_data_nota.c.dataDigitacaoNotaFiscal, ContaPagar.Data_Digitacao)
        valor_efetivo = func.coalesce(
            func.nullif(ContaPagar.Valor_RecebidoNotaFiscal, 0),
            ContaPagar.Valor_ContaPagar,
            0,
        )

        registros = (
            sessao.query(
                extract('month', data_efetiva).label('mesCompetencia'),
                PlanoConta.Numero_ContaContabil.label('numeroContaContabil'),
                PlanoConta.Descricao_ContaContabil.label('descricaoContaContabil'),
                Fornecedor.Nome_Fornecedor.label('nomeFornecedor'),
                ContaPagar.Descricao_Item.label('descricaoItem'),
                ContaPagar.Numero_Documento.label('numeroDocumento'),
                ContaPagar.Opcao_StatusContaPagar.label('codigoStatusContaPagar'),
                ContaPagar.Data_Emissao.label('dataEmissao'),
                valor_efetivo.label('valor'),
            )
            .select_from(ContaPagar)
            .outerjoin(
                subconsulta_data_nota,
                ContaPagar.Codigo_ContaPagar == subconsulta_data_nota.c.codigoContaPagar,
            )
            .outerjoin(PlanoConta, ContaPagar.Codigo_ContaContabil == PlanoConta.Codigo_ContaContabil)
            .outerjoin(Fornecedor, ContaPagar.Codigo_Fornecedor == Fornecedor.Codigo_Fornecedor)
            .filter(ContaPagar.Codigo_CentroCusto == codigo_centro_custo)
            .filter(extract('year', data_efetiva) == ano)
            .filter(ContaPagar.Opcao_StatusContaPagar.in_(self.STATUS_CONSIDERADOS))
            .filter(valor_efetivo != 0)
            .order_by(
                extract('month', data_efetiva),
                data_efetiva,
                ContaPagar.Data_Emissao,
                PlanoConta.Numero_ContaContabil,
                Fornecedor.Nome_Fornecedor,
            )
        )

        if codigos_conta_contabil:
            registros = registros.filter(ContaPagar.Codigo_ContaContabil.in_(codigos_conta_contabil))

        registros = registros.all()

        lancamentos_por_mes = {
            numero: []
            for numero, _, _, _ in self.MESES
        }

        for registro in registros:
            numero_mes = int(registro.mesCompetencia or 0)
            if numero_mes not in lancamentos_por_mes:
                continue

            conta = self._montarTextoConta(registro.numeroContaContabil, registro.descricaoContaContabil)
            fornecedor = self._normalizarTexto(registro.nomeFornecedor) or 'Sem fornecedor vinculado'
            numero_documento = self._normalizarTexto(registro.numeroDocumento) or 'Sem documento informado'
            status = self._descreverStatusContaPagar(registro.codigoStatusContaPagar)
            descricao = self._normalizarTexto(registro.descricaoItem) or 'Sem descricao'

            lancamentos_por_mes[numero_mes].append(
                {
                    'conta_contabil': conta,
                    'fornecedor': fornecedor,
                    'numero_documento': numero_documento,
                    'status': status,
                    'descricao': descricao,
                    'data_emissao': registro.dataEmissao,
                    'valor': float(registro.valor or 0.0),
                }
            )

        return lancamentos_por_mes

    def _montarTextoConta(self, numero, descricao):
        descricao_texto = self._normalizarTexto(descricao)
        numero_texto = self._normalizarTexto(numero)
        return descricao_texto or numero_texto or 'Sem conta contabil vinculada'

    def _descreverStatusContaPagar(self, codigo_status):
        try:
            status_normalizado = int(codigo_status)
        except (TypeError, ValueError):
            return 'Status nao mapeado'

        if status_normalizado in self.STATUS_EM_APROVACAO:
            return 'Em Aprovação'

        if status_normalizado in self.STATUS_APROVADO:
            return 'Aprovado'

        return 'Status nao mapeado'

    def _atualizarAbaLista(self, planilha, ano, centro):
        for indice, (_, _, abreviacao, _) in enumerate(self.MESES, start=2):
            planilha.cell(row=indice, column=2).value = f'{abreviacao}/{ano}'

        for indice in range(2, 200):
            planilha.cell(row=indice, column=4).value = None
            planilha.cell(row=indice, column=5).value = None

        planilha.cell(row=2, column=4).value = self._normalizarNumero(centro['codigo'])
        planilha.cell(row=2, column=5).value = centro['nome']

    def _configurarCabecalhosAbaMensal(self, planilha):
        planilha.cell(row=17, column=5).value = 'NÚMERO DOCUMENTO'
        planilha.cell(row=17, column=6).value = 'STATUS'
        planilha.cell(row=17, column=7).value = 'DESCRIÇÃO'
        planilha.cell(row=17, column=8).value = 'DATA DE EMISSÃO'
        planilha.cell(row=17, column=9)._style = copy(planilha.cell(row=17, column=8)._style)
        planilha.cell(row=17, column=9).value = 'VALOR'

        planilha.column_dimensions['E'].width = 18
        planilha.column_dimensions['F'].width = 16
        planilha.column_dimensions['G'].width = 30
        planilha.column_dimensions['H'].width = 15
        planilha.column_dimensions['I'].width = 14

    def _limparCelulaSemFormato(self, planilha, numero_linha, indice_coluna, coluna_referencia=10):
        celula = planilha.cell(row=numero_linha, column=indice_coluna)
        celula.value = None
        celula._style = copy(planilha.cell(row=numero_linha, column=coluna_referencia)._style)

    def _configurarResumoSuperior(self, planilha):
        estilo_rotulo_11 = copy(planilha.cell(row=11, column=7)._style)
        estilo_valor_11 = copy(planilha.cell(row=11, column=8)._style)
        planilha.cell(row=11, column=8)._style = copy(estilo_rotulo_11)
        planilha.cell(row=11, column=8).value = 'BUDGET/ANO'
        planilha.cell(row=11, column=9)._style = copy(estilo_valor_11)
        self._limparCelulaSemFormato(planilha, 11, 7)

        estilo_rotulo_13 = copy(planilha.cell(row=13, column=7)._style)
        estilo_valor_13 = copy(planilha.cell(row=13, column=8)._style)
        planilha.cell(row=13, column=8)._style = copy(estilo_rotulo_13)
        planilha.cell(row=13, column=8).value = 'BUDGET/MÊS'
        planilha.cell(row=13, column=9)._style = copy(estilo_valor_13)
        self._limparCelulaSemFormato(planilha, 13, 7)

        if 'F15:G15' in [str(faixa) for faixa in planilha.merged_cells.ranges]:
            planilha.unmerge_cells('F15:G15')

        estilo_rotulo_15 = copy(planilha.cell(row=15, column=6)._style)
        estilo_valor_15 = copy(planilha.cell(row=15, column=8)._style)
        planilha.cell(row=15, column=7)._style = copy(estilo_rotulo_15)
        planilha.cell(row=15, column=7).value = 'CONSUMO ACUMULADO ATÉ O MÊS ANTERIOR'
        planilha.cell(row=15, column=8)._style = copy(estilo_rotulo_15)
        self._limparCelulaSemFormato(planilha, 15, 6)
        planilha.cell(row=15, column=9)._style = copy(estilo_valor_15)
        planilha.merge_cells(start_row=15, start_column=7, end_row=15, end_column=8)

    def _configurarLinhasResumo(
        self,
        planilha,
        linha_total,
        linha_saldo_mensal,
        linha_percentual_mensal,
        linha_acumulado,
        linha_saldo_acumulado,
        linha_percentual_ano,
    ):
        linhas_resumo = (
            (linha_total, 'TOTAL DO MÊS'),
            (linha_saldo_mensal, 'SALDO DO BUDGET MENSAL'),
            (linha_percentual_mensal, '%BUDGET CONSUMIDO / MÊS'),
            (linha_acumulado, 'CONSUMO ACUMULADO ATUAL'),
            (linha_saldo_acumulado, 'SALDO DO BUDGET ACUMULADO'),
            (linha_percentual_ano, '% BUDGET CONSUMIDO / ANO'),
        )

        for numero_linha, rotulo in linhas_resumo:
            estilo_rotulo = copy(planilha.cell(row=numero_linha, column=6)._style)
            self._limparCelulaSemFormato(planilha, numero_linha, 6)
            planilha.cell(row=numero_linha, column=7)._style = copy(estilo_rotulo)
            planilha.cell(row=numero_linha, column=7).value = rotulo
            planilha.cell(row=numero_linha, column=8).value = None
            planilha.cell(row=numero_linha, column=9)._style = copy(planilha.cell(row=numero_linha, column=8)._style)

    def _preencherAbaMensal(
        self,
        planilha,
        ano,
        numero_mes,
        abreviacao_mes,
        gestor,
        centro,
        total_ano,
        total_mes,
        lancamentos,
        nome_aba_anterior,
        linha_acumulada_anterior,
    ):
        self._configurarResumoSuperior(planilha)
        self._configurarCabecalhosAbaMensal(planilha)
        modelos = self._capturarModelosLinha(planilha)
        quantidade_linhas = max(self.QUANTIDADE_MINIMA_LINHAS, len(lancamentos))

        linha_total = self.LINHA_INICIAL_DETALHE + quantidade_linhas
        linha_saldo_mensal = linha_total + 2
        linha_percentual_mensal = linha_total + 3
        linha_acumulado = linha_total + 6
        linha_saldo_acumulado = linha_total + 7
        linha_percentual_ano = linha_total + 8
        linha_aprovador = linha_total + 11
        linha_data = linha_total + 12
        linha_final = linha_data

        self._limparRegiaoDinamica(planilha, max(planilha.max_row, linha_final))

        self._aplicarModeloLinha(planilha, self.LINHA_INTRO_DETALHE, modelos['intro'])
        for deslocamento in range(quantidade_linhas):
            linha_atual = self.LINHA_INICIAL_DETALHE + deslocamento
            self._aplicarModeloLinha(planilha, linha_atual, modelos['detalhe'])

        self._aplicarModeloLinha(planilha, linha_total, modelos['total'])
        self._aplicarModeloLinha(planilha, linha_total + 1, modelos['blank_1'])
        self._aplicarModeloLinha(planilha, linha_saldo_mensal, modelos['saldo_mensal'])
        self._aplicarModeloLinha(planilha, linha_percentual_mensal, modelos['percentual_mensal'])
        self._aplicarModeloLinha(planilha, linha_percentual_mensal + 1, modelos['blank_2'])
        self._aplicarModeloLinha(planilha, linha_percentual_mensal + 2, modelos['blank_3'])
        self._aplicarModeloLinha(planilha, linha_acumulado, modelos['acumulado'])
        self._aplicarModeloLinha(planilha, linha_saldo_acumulado, modelos['saldo_acumulado'])
        self._aplicarModeloLinha(planilha, linha_percentual_ano, modelos['percentual_ano'])
        self._aplicarModeloLinha(planilha, linha_percentual_ano + 1, modelos['blank_4'])
        self._aplicarModeloLinha(planilha, linha_percentual_ano + 2, modelos['blank_5'])
        self._aplicarModeloLinha(planilha, linha_aprovador, modelos['aprovador'])
        self._aplicarModeloLinha(planilha, linha_data, modelos['data'])

        self._configurarLinhasResumo(
            planilha,
            linha_total=linha_total,
            linha_saldo_mensal=linha_saldo_mensal,
            linha_percentual_mensal=linha_percentual_mensal,
            linha_acumulado=linha_acumulado,
            linha_saldo_acumulado=linha_saldo_acumulado,
            linha_percentual_ano=linha_percentual_ano,
        )
        self._aplicarMesclagens(
            planilha,
            quantidade_linhas=quantidade_linhas,
            linha_total=linha_total,
            linha_saldo_mensal=linha_saldo_mensal,
            linha_percentual_ano=linha_percentual_ano,
            linha_aprovador=linha_aprovador,
            linha_data=linha_data,
        )

        planilha['D1'] = f'ANÁLISE  MENSAL  BUDGET {ano}'
        planilha['C6'] = f'{abreviacao_mes}/{ano}'
        planilha['C7'] = self._montarResponsavel(gestor)
        planilha['C8'] = f"{centro['nome']}"
        planilha['I11'] = total_ano
        planilha['I13'] = total_mes
        planilha['I11'].number_format = self.FORMATO_MOEDA
        planilha['I13'].number_format = self.FORMATO_MOEDA

        if numero_mes == 1:
            planilha['I15'] = 0
        else:
            planilha['I15'] = f"='{nome_aba_anterior}'!I{linha_acumulada_anterior}"
        planilha['I15'].number_format = self.FORMATO_MOEDA

        for indice in range(quantidade_linhas):
            linha_atual = self.LINHA_INICIAL_DETALHE + indice
            lancamento = lancamentos[indice] if indice < len(lancamentos) else None

            estilo_texto = copy(planilha.cell(row=linha_atual, column=5)._style)
            estilo_data = copy(planilha.cell(row=linha_atual, column=6)._style)
            estilo_moeda = copy(planilha.cell(row=linha_atual, column=8)._style)
            planilha.cell(row=linha_atual, column=6)._style = copy(estilo_texto)
            planilha.cell(row=linha_atual, column=7)._style = copy(estilo_texto)
            planilha.cell(row=linha_atual, column=8)._style = copy(estilo_data)
            planilha.cell(row=linha_atual, column=9)._style = copy(estilo_moeda)

            planilha.cell(row=linha_atual, column=2).value = lancamento['conta_contabil'] if lancamento else None
            planilha.cell(row=linha_atual, column=3).value = lancamento['fornecedor'] if lancamento else None
            planilha.cell(row=linha_atual, column=5).value = lancamento['numero_documento'] if lancamento else None
            planilha.cell(row=linha_atual, column=6).value = lancamento['status'] if lancamento else None
            planilha.cell(row=linha_atual, column=7).value = lancamento['descricao'] if lancamento else None
            planilha.cell(row=linha_atual, column=8).value = lancamento['data_emissao'] if lancamento else None
            planilha.cell(row=linha_atual, column=9).value = lancamento['valor'] if lancamento else None
            planilha.cell(row=linha_atual, column=8).number_format = self.FORMATO_DATA
            planilha.cell(row=linha_atual, column=9).number_format = self.FORMATO_MOEDA

        planilha.cell(row=linha_total, column=9).value = (
            f'=SUM(I{self.LINHA_INICIAL_DETALHE}:I{linha_total - 1})'
        )
        planilha.cell(row=linha_total, column=9).number_format = self.FORMATO_MOEDA
        planilha.cell(row=linha_saldo_mensal, column=9).value = f'=IF(I13=0,0,I13-I{linha_total})'
        planilha.cell(row=linha_saldo_mensal, column=9).number_format = self.FORMATO_MOEDA
        planilha.cell(row=linha_percentual_mensal, column=9).value = f'=IF(I13=0,0,I{linha_total}/I13)'
        planilha.cell(row=linha_percentual_mensal, column=9).number_format = '0.00%'
        planilha.cell(row=linha_acumulado, column=9).value = f'=I15+I{linha_total}'
        planilha.cell(row=linha_acumulado, column=9).number_format = self.FORMATO_MOEDA
        planilha.cell(row=linha_saldo_acumulado, column=9).value = f'=I11-I{linha_acumulado}'
        planilha.cell(row=linha_saldo_acumulado, column=9).number_format = self.FORMATO_MOEDA
        planilha.cell(row=linha_percentual_ano, column=9).value = f'=IFERROR(I{linha_acumulado}/I11,0)'
        planilha.cell(row=linha_percentual_ano, column=9).number_format = '0.00%'

        return linha_acumulado

    def _capturarModelosLinha(self, planilha):
        return {
            'intro': self._capturarLinha(planilha, 18),
            'detalhe': self._capturarLinha(planilha, 19),
            'total': self._capturarLinha(planilha, 28),
            'blank_1': self._capturarLinha(planilha, 29),
            'saldo_mensal': self._capturarLinha(planilha, 30),
            'percentual_mensal': self._capturarLinha(planilha, 31),
            'blank_2': self._capturarLinha(planilha, 32),
            'blank_3': self._capturarLinha(planilha, 33),
            'acumulado': self._capturarLinha(planilha, 34),
            'saldo_acumulado': self._capturarLinha(planilha, 35),
            'percentual_ano': self._capturarLinha(planilha, 36),
            'blank_4': self._capturarLinha(planilha, 37),
            'blank_5': self._capturarLinha(planilha, 38),
            'aprovador': self._capturarLinha(planilha, 39),
            'data': self._capturarLinha(planilha, 40),
        }

    def _capturarLinha(self, planilha, numero_linha):
        return {
            'altura': planilha.row_dimensions[numero_linha].height,
            'colunas': {
                indice_coluna: {
                    'valor': planilha.cell(row=numero_linha, column=indice_coluna).value,
                    'estilo': copy(planilha.cell(row=numero_linha, column=indice_coluna)._style),
                }
                for indice_coluna in self.COLUNAS_PLANILHA
            },
        }

    def _aplicarModeloLinha(self, planilha, numero_linha, modelo):
        planilha.row_dimensions[numero_linha].height = modelo['altura']
        for indice_coluna, configuracao in modelo['colunas'].items():
            celula = planilha.cell(row=numero_linha, column=indice_coluna)
            celula._style = copy(configuracao['estilo'])
            celula.value = configuracao['valor']

    def _limparRegiaoDinamica(self, planilha, ultima_linha):
        for faixa_mesclada in list(planilha.merged_cells.ranges):
            if faixa_mesclada.max_row >= self.LINHA_INTRO_DETALHE and faixa_mesclada.min_row <= ultima_linha:
                planilha.unmerge_cells(str(faixa_mesclada))

        for numero_linha in range(self.LINHA_INTRO_DETALHE, ultima_linha + 1):
            for indice_coluna in self.COLUNAS_PLANILHA:
                planilha.cell(row=numero_linha, column=indice_coluna).value = None

    def _aplicarMesclagens(
        self,
        planilha,
        quantidade_linhas,
        linha_total,
        linha_saldo_mensal,
        linha_percentual_ano,
        linha_aprovador,
        linha_data,
    ):
        planilha.merge_cells(start_row=18, start_column=2, end_row=18, end_column=9)

        for deslocamento in range(quantidade_linhas):
            linha_atual = self.LINHA_INICIAL_DETALHE + deslocamento
            planilha.merge_cells(start_row=linha_atual, start_column=3, end_row=linha_atual, end_column=4)

        planilha.merge_cells(start_row=linha_total, start_column=7, end_row=linha_total, end_column=8)
        planilha.merge_cells(start_row=linha_saldo_mensal, start_column=7, end_row=linha_saldo_mensal, end_column=8)
        planilha.merge_cells(start_row=linha_saldo_mensal + 1, start_column=7, end_row=linha_saldo_mensal + 1, end_column=8)
        planilha.merge_cells(start_row=linha_saldo_mensal + 4, start_column=7, end_row=linha_saldo_mensal + 4, end_column=8)
        planilha.merge_cells(start_row=linha_saldo_mensal + 5, start_column=7, end_row=linha_saldo_mensal + 5, end_column=8)
        planilha.merge_cells(start_row=linha_percentual_ano, start_column=2, end_row=linha_percentual_ano, end_column=3)
        planilha.merge_cells(start_row=linha_percentual_ano, start_column=7, end_row=linha_percentual_ano, end_column=8)
        planilha.merge_cells(start_row=linha_percentual_ano + 1, start_column=2, end_row=linha_percentual_ano + 1, end_column=3)
        planilha.merge_cells(start_row=linha_aprovador, start_column=2, end_row=linha_aprovador, end_column=3)
        planilha.merge_cells(start_row=linha_data, start_column=2, end_row=linha_data, end_column=3)

    def _montarResponsavel(self, gestor):
        nome = self._normalizarTexto(gestor.get('nome_usuario')) or 'Gestor'
        
        """
        - Comentado para evitar que o cargo do gestor seja exibido na planilha de 
        acompanhamento mensal,

        
        cargo = self._normalizarTexto(gestor.get('cargo'))
        if cargo:
            return f'{nome} - {cargo}'
        return nome
        """
        return nome

    def _normalizarTexto(self, valor):
        if valor is None:
            return None

        texto = str(valor).strip()
        return texto or None

    def _normalizarNumero(self, valor):
        texto = self._normalizarTexto(valor)
        if texto is None:
            return None

        return int(texto) if texto.isdigit() else texto

    def _normalizarNomeArquivo(self, valor):
        texto = self._normalizarTexto(valor) or 'arquivo'
        permitido = [caractere if caractere.isalnum() else '_' for caractere in texto]
        return ''.join(permitido).strip('_') or 'arquivo'