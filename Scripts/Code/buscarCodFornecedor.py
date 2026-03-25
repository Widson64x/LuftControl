import os
import sys
import re
import shutil
import pandas as pd
import openpyxl
from sqlalchemy import text

# Configuração do caminho do sistema para importação do módulo Db
diretorioAtual = os.path.dirname(os.path.abspath(__file__))
diretorioRaiz = os.path.abspath(os.path.join(diretorioAtual, '..', '..'))
if diretorioRaiz not in sys.path:
    sys.path.append(diretorioRaiz)

try:
    from Db.Connections import GetSqlServerEngine
except ImportError as erroImportacao:
    print(f"Erro Crítico: Não foi possível importar o módulo Db.Connections. Detalhes: {erroImportacao}")
    sys.exit(1)


def BuscarCodigoFornecedor(conexaoSql, descricaoFornecedor):
    """
    Realiza uma consulta no banco de dados com múltiplas tentativas progressivas (Fallback) 
    para localizar o código do fornecedor, contornando abreviações, traços, vírgulas ou parênteses.
    
    Parâmetros:
        conexaoSql (sqlalchemy.engine.Connection): Instância de conexão ativa.
        descricaoFornecedor (str): String contendo o nome ou descrição do fornecedor.
        
    Retornos:
        str ou None: Retorna o identificador do código do fornecedor se encontrado, 
        caso contrário retorna None.
    """
    if pd.isna(descricaoFornecedor) or not str(descricaoFornecedor).strip():
        return None
        
    descricaoLimpa = str(descricaoFornecedor).strip()
    
    def ExecutarConsultaInterna(termo):
        """Função auxiliar para executar a query no banco evitando repetição de código."""
        query = text("SELECT Codigo_Fornecedor FROM Fornecedor WHERE Nome_Fornecedor LIKE :termo")
        resultadoQuery = conexaoSql.execute(query, {"termo": termo}).fetchone()
        return resultadoQuery[0] if resultadoQuery else None

    codigo = ExecutarConsultaInterna(f"%{descricaoLimpa}%")
    if codigo: return codigo

    descricaoSemSufixo = re.split(r'\(|-|,', descricaoLimpa)[0].strip()
    if descricaoSemSufixo and descricaoSemSufixo != descricaoLimpa:
        codigo = ExecutarConsultaInterna(f"%{descricaoSemSufixo}%")
        if codigo: return codigo

    palavras = descricaoSemSufixo.split()
    if len(palavras) >= 2:
        termoDuasPalavras = f"%{palavras[0]}%{palavras[1]}%"
        codigo = ExecutarConsultaInterna(termoDuasPalavras)
        if codigo: return codigo

    if len(palavras) >= 1 and len(palavras[0]) > 4:
        termoPrimeiraPalavra = f"%{palavras[0]}%"
        codigo = ExecutarConsultaInterna(termoPrimeiraPalavra)
        if codigo: return codigo

    return None


def ProcessarPlanilhaFornecedores(caminhoArquivoEntrada, caminhoArquivoSaida, nomeAba):
    """
    Lê o Excel para identificar fornecedores faltantes, busca no banco de dados e 
    atualiza o arquivo copiando-o e injetando os dados diretamente nas células 
    via openpyxl, preservando assim toda a formatação (cores, bordas, etc).
    
    Parâmetros:
        caminhoArquivoEntrada (str): Caminho local do arquivo Excel base.
        caminhoArquivoSaida (str): Caminho local onde o Excel atualizado será salvo.
        nomeAba (str): Nome da aba dentro do arquivo Excel.
        
    Retornos:
        None
    """
    engineSql = GetSqlServerEngine()
    if engineSql is None:
        print("Erro: A função GetSqlServerEngine retornou None. Verifique as conexões.")
        return

    # Utiliza o pandas apenas para leitura rápida e identificação dos valores únicos
    try:
        dataframeLeitura = pd.read_excel(caminhoArquivoEntrada, sheet_name=nomeAba)
    except ValueError as erroLeitura:
        print(f"Erro ao ler a planilha: {erroLeitura}.")
        return

    dataframeLeitura.columns = dataframeLeitura.columns.str.strip()
    
    if 'DescFornecedor' not in dataframeLeitura.columns or 'CodFornecedor' not in dataframeLeitura.columns:
        print(f"Erro: As colunas necessárias não estão presentes na aba '{nomeAba}'.")
        return

    # Isola os fornecedores que estão com o código vazio
    mascaraVazios = dataframeLeitura['CodFornecedor'].isna() | (dataframeLeitura['CodFornecedor'].astype(str).str.strip() == "")
    fornecedoresUnicos = dataframeLeitura.loc[mascaraVazios, 'DescFornecedor'].dropna().unique()
    
    dicionarioCodigos = {}
    
    print(f"Iniciando busca inteligente no banco de dados para {len(fornecedoresUnicos)} fornecedores únicos...")

    with engineSql.connect() as conexaoSql:
        for descricao in fornecedoresUnicos:
            codigoBusca = BuscarCodigoFornecedor(conexaoSql, descricao)
            if codigoBusca is not None:
                dicionarioCodigos[descricao] = codigoBusca
                print(f"DEBUG: SUCESSO - '{descricao}' -> Código: {codigoBusca}")
            else:
                dicionarioCodigos[descricao] = None
                print(f"DEBUG: FALHA   - '{descricao}' não localizado.")

    print("\nCopiando o arquivo original para preservar formatação e cores...")
    try:
        shutil.copy2(caminhoArquivoEntrada, caminhoArquivoSaida)
    except Exception as erroCopia:
        print(f"Erro ao criar cópia do arquivo original: {erroCopia}")
        return

    print("Injetando os códigos atualizados diretamente nas células...")
    
    try:
        # Carrega a pasta de trabalho utilizando openpyxl para manipulação a nível de célula
        pastaTrabalho = openpyxl.load_workbook(caminhoArquivoSaida)
        planilhaBase = pastaTrabalho[nomeAba]
        
        indiceColunaCodigo = None
        indiceColunaDescricao = None
        
        # Identifica dinamicamente as colunas com base no cabeçalho (linha 1)
        for celula in planilhaBase[1]:
            if celula.value and str(celula.value).strip() == 'CodFornecedor':
                indiceColunaCodigo = celula.column
            elif celula.value and str(celula.value).strip() == 'DescFornecedor':
                indiceColunaDescricao = celula.column
                
        if not indiceColunaCodigo or not indiceColunaDescricao:
            print("Erro: Não foi possível localizar os índices das colunas no openpyxl.")
            return

        listaNaoEncontrados = []

        # Itera pelas linhas a partir da segunda (ignorando o cabeçalho)
        for linhaAtual in range(2, planilhaBase.max_row + 1):
            celulaCodigo = planilhaBase.cell(row=linhaAtual, column=indiceColunaCodigo)
            celulaDescricao = planilhaBase.cell(row=linhaAtual, column=indiceColunaDescricao)
            
            valorCodigo = celulaCodigo.value
            valorDescricao = celulaDescricao.value
            
            # Se a célula de código estiver vazia, tenta preencher baseando-se no dicionário
            if valorCodigo is None or str(valorCodigo).strip() == "":
                if valorDescricao in dicionarioCodigos and dicionarioCodigos[valorDescricao] is not None:
                    # Atualiza o valor mantendo a formatação existente da célula
                    celulaCodigo.value = dicionarioCodigos[valorDescricao]
                elif valorDescricao:
                    listaNaoEncontrados.append(valorDescricao)

        # Trata a aba de não encontrados
        nomeAbaNaoEncontrados = 'Nao_Encontrados'
        if nomeAbaNaoEncontrados in pastaTrabalho.sheetnames:
            del pastaTrabalho[nomeAbaNaoEncontrados]
            
        planilhaNaoEncontrados = pastaTrabalho.create_sheet(title=nomeAbaNaoEncontrados)
        planilhaNaoEncontrados.append(['Fornecedores_Nao_Localizados'])
        
        if listaNaoEncontrados:
            fornecedoresUnicosFalhos = list(set(listaNaoEncontrados))
            for item in fornecedoresUnicosFalhos:
                planilhaNaoEncontrados.append([item])
        else:
            planilhaNaoEncontrados.append(['Todos os fornecedores foram encontrados com sucesso.'])

        # Salva as alterações estruturais no arquivo físico
        pastaTrabalho.save(caminhoArquivoSaida)
        print(f"Processamento concluído. O arquivo atualizado preservando as cores foi salvo em: {caminhoArquivoSaida}")
        
    except Exception as erroEdicao:
        print(f"Erro ao editar células no arquivo Excel: {erroEdicao}")


if __name__ == "__main__":
    arquivoEntrada = r"C:\Applications\Python\Projetos\LuftControl\Data\CodFornecedor.xlsx"
    arquivoSaida = r"C:\Applications\Python\Projetos\LuftControl\Data\CodFornecedor_Atualizado.xlsx"
    nomeAbaDados = "Base"
    
    ProcessarPlanilhaFornecedores(arquivoEntrada, arquivoSaida, nomeAbaDados)